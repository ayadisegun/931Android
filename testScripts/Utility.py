import openpyxl
import pytest


def read_excel_data(file_path, sheet_name):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
    except FileNotFoundError:
        pytest.fail(f"Excel file not found: {file_path}")
    except KeyError:
        pytest.fail(f"Sheet '{sheet_name}' not found in the Excel file")

    max_row = sheet.max_row
    max_col = sheet.max_column

    # Get headers from first row
    headers = [sheet.cell(row=1, column=col).value for col in range(1, max_col + 1)]
    print("headers are: ", headers)

    data_rows = []
    for row in range(2, max_row + 1):
        row_data = [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
        data_rows.append(tuple(row_data))

    return headers, data_rows


def get_Excel_data_path_in_script():
    # return [
    #
    #     ["segun", "Ayadi", "Rapture", "segema@gmail", "Testing1", "techsupport@creditswitch.com", "Testing2."],
    #     ["taye", "taiwo", "tailoo", "taiwo@gmail", "Testing4", "ayadisegun02@gmail.com", "Credit2."],
    #
    # ]
    excelfile = "C:\\Users\\Segun\\PycharmProjects\\Android931\\testScripts\\ExcelFile.xlsx"
    sheetName = "details"
    workbook = openpyxl.load_workbook(excelfile)
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2, totalrows+1):
        dataList = []
        for j in range(1, totalcols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList


def setCellData_path_in_script(rowNum, colNum, data):
    excelfile = "C:\\Users\\Segun\\PycharmProjects\\Android931\\testScripts\\ExcelFile.xlsx"
    sheetName = "details"
    workbook = openpyxl.load_workbook(excelfile)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value=data
    workbook.save(excelfile)



#
# excelfile = "C:\\Users\\Segun\\PycharmProjects\\Android931\\testScripts\\ExcelFile.xlsx"
# page = "details"
# workbook = openpyxl.load_workbook(excelfile)
# sheet = workbook[page]
# totalrows =  sheet.max_row
# totalcols = sheet.max_column
#
#
# def getRowCount(path, sheetName):
#     workbook = openpyxl.load_workbook(excelfile)
#     sheet = workbook[page]
#     return sheet.max_row
#
#
# def getColCount(path, sheetName):
#     workbook = openpyxl.load_workbook(excelfile)
#     sheet = workbook[page]
#     return sheet.max_column
#
#
# def getCellData(path, sheetName, rowNum, colNum):
#     workbook = openpyxl.load_workbook(excelfile)
#     sheet = workbook[page]
#     return sheet.cell(row=rowNum, column=colNum).value
#
#
# def setCellData1(path, sheetName, rowNum, colNum, data):
#     workbook = openpyxl.load_workbook(excelfile)
#     sheet = workbook[page]
#     sheet.cell(row=rowNum, column=colNum).value=data
#     workbook.save(path)


# rows = getRowCount(excelfile, page)
# print(rows)
# columns = getColCount(excelfile, page)
# print(columns)
# firstUser = getCellData(excelfile,page,2,1)
# print("firstUser is: ", firstUser)
# setCellData1(excelfile,page,2,1, "segun")
# firstUpdated = getCellData(excelfile,page,2,1)
# print("firstUser is updated to: ", firstUpdated)

